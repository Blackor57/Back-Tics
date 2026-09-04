# app/services/excel_reporter.py
from pathlib import Path
from typing import Dict, Any, List, Optional
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage


class ExcelReporter:
    @staticmethod
    def generar_reporte(
        url: str,
        site_title: str,
        analisis_ia: Dict[str, Any],
        articulos: List[Dict[str, Any]],
        delta: Optional[Dict[str, Any]],
        chart_paths: List[Path],
        output_file: Path
    ) -> Path:
        """
        Crea un libro Excel (.xlsx) estructurado en 3 hojas especializadas:
        1. Dashboard Ejecutivo & KPIs
        2. Detalle de Contenidos (con auto-filtros y estados)
        3. Matriz de Inteligencia (Categorías, Entidades y Recomendaciones)
        """
        wb = openpyxl.Workbook()

        # Paleta de estilos corporativa
        font_title = Font(name="Segoe UI", size=15, bold=True, color="0F172A")
        font_sub = Font(name="Segoe UI", size=9.5, color="64748B")
        font_sec = Font(name="Segoe UI", size=11, bold=True, color="1E3A8A")
        font_header = Font(name="Segoe UI", size=9.5, bold=True, color="FFFFFF")
        font_bold = Font(name="Segoe UI", size=9, bold=True, color="0F172A")
        font_normal = Font(name="Segoe UI", size=9, color="334155")

        fill_navy = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        fill_slate = PatternFill(start_color="334155", end_color="334155", fill_type="solid")
        fill_zebra = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        fill_novedad = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
        fill_card_bg = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")

        border_thin = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0')
        )

        total_items = len(articulos) if isinstance(articulos, list) else 1
        total_nuevos = delta.get("total_nuevos", 0) if delta else 0
        rotacion = delta.get("tasa_rotacion_pct", 0) if delta else 0.0
        nivel_alerta = str(analisis_ia.get("nivel_alerta", "MEDIO")).upper()

        # =========================================================
        # HOJA 1: DASHBOARD EJECUTIVO
        # =========================================================
        ws1 = wb.active
        ws1.title = "Dashboard Ejecutivo"
        ws1.views.sheetView[0].showGridLines = True

        # Encabezado
        ws1["A1"] = "INFORME DE AUDITORÍA Y MONITOREO WEB"
        ws1["A1"].font = font_title
        ws1["A2"] = f"Objetivo: {site_title} | URL: {url}"
        ws1["A2"].font = font_sub
        ws1["A3"] = f"Fecha de Emisión: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')} | Tipo: {analisis_ia.get('tipo_portal', 'Portal Web')}"
        ws1["A3"].font = font_sub

        # Tarjetas de KPIs (Fila 5 a 6)
        kpis = [
            ("A", "B", "TOTAL ELEMENTOS", str(total_items)),
            ("C", "D", "NOVEDADES DETECTADAS", f"+{total_nuevos}"),
            ("E", "F", "TASA DE ROTACIÓN", f"{rotacion}%"),
            ("G", "H", "NIVEL DE ALERTA", nivel_alerta)
        ]

        for col1, col2, titulo_kpi, val_kpi in kpis:
            ws1.merge_cells(f"{col1}5:{col2}5")
            ws1.merge_cells(f"{col1}6:{col2}6")
            ws1[f"{col1}5"] = titulo_kpi
            ws1[f"{col1}5"].font = Font(name="Segoe UI", size=8, bold=True, color="64748B")
            ws1[f"{col1}5"].alignment = Alignment(horizontal="center", vertical="center")
            ws1[f"{col1}5"].fill = fill_card_bg

            ws1[f"{col1}6"] = val_kpi
            ws1[f"{col1}6"].font = Font(name="Segoe UI", size=14, bold=True, color="0F172A")
            ws1[f"{col1}6"].alignment = Alignment(horizontal="center", vertical="center")
            ws1[f"{col1}6"].fill = fill_card_bg

            for r in [5, 6]:
                for c_letter in [col1, col2]:
                    ws1[f"{c_letter}{r}"].border = border_thin

        # Resumen Ejecutivo
        ws1["A8"] = "Resumen Ejecutivo de Inteligencia"
        ws1["A8"].font = font_sec
        ws1.merge_cells("A9:D13")
        ws1["A9"] = analisis_ia.get("resumen_ejecutivo", "")
        ws1["A9"].font = font_normal
        ws1["A9"].alignment = Alignment(wrap_text=True, vertical="top")
        for r in range(9, 14):
            for c_letter in ["A", "B", "C", "D"]:
                ws1[f"{c_letter}{r}"].fill = fill_zebra
                ws1[f"{c_letter}{r}"].border = border_thin

        # Puntos de Atención Urgentes
        ws1["A15"] = "Puntos de Atención Urgentes y Conclusiones"
        ws1["A15"].font = font_sec
        puntos = analisis_ia.get("puntos_atencion_urgentes", []) + analisis_ia.get("conclusiones", [])
        ws1.merge_cells("A16:D20")
        ws1["A16"] = "\n".join([f"• {p}" for p in puntos[:6]])
        ws1["A16"].font = font_normal
        ws1["A16"].alignment = Alignment(wrap_text=True, vertical="top")
        for r in range(16, 21):
            for c_letter in ["A", "B", "C", "D"]:
                ws1[f"{c_letter}{r}"].fill = fill_zebra
                ws1[f"{c_letter}{r}"].border = border_thin

        # Incrustar Gráficos en Hoja 1
        img_anchors = ["E8", "E22", "A22"]
        for idx, cp in enumerate(chart_paths[:3]):
            if cp.exists():
                try:
                    img = OpenpyxlImage(str(cp))
                    img.width = 440
                    img.height = 250
                    ws1.add_image(img, img_anchors[idx])
                except Exception:
                    pass

        # Anchos de columna en Hoja 1
        for col_name, w in [("A", 18), ("B", 20), ("C", 18), ("D", 20), ("E", 25), ("F", 25), ("G", 22), ("H", 22)]:
            ws1.column_dimensions[col_name].width = w

        # =========================================================
        # HOJA 2: DETALLE DE CONTENIDOS (CON AUTO-FILTRO)
        # =========================================================
        ws2 = wb.create_sheet(title="Detalle de Contenidos")
        ws2.views.sheetView[0].showGridLines = True

        headers_det = ["N°", "Título / Asunto Detectado", "Estado", "Caracteres", "Enlace Web Directo"]
        for col_idx, h in enumerate(headers_det, start=1):
            cell = ws2.cell(row=1, column=col_idx)
            cell.value = h
            cell.font = font_header
            cell.fill = fill_navy
            cell.alignment = Alignment(horizontal="center", vertical="center")

        nuevas_urls = set()
        if delta and delta.get("nuevos_articulos"):
            nuevas_urls = {item.get("url") for item in delta.get("nuevos_articulos") if isinstance(item, dict)}

        lista_articulos = articulos if isinstance(articulos, list) else [{"titulo": site_title, "url": url}]
        for row_idx, art in enumerate(lista_articulos, start=2):
            art_url = art.get("url", "")
            es_novedad = art_url in nuevas_urls
            fill_actual = fill_novedad if es_novedad else (fill_zebra if row_idx % 2 == 0 else None)

            ws2.cell(row=row_idx, column=1, value=row_idx - 1).alignment = Alignment(horizontal="center")
            ws2.cell(row=row_idx, column=2, value=art.get("titulo", "Sin título")[:180])
            cell_est = ws2.cell(row=row_idx, column=3, value="Novedad" if es_novedad else "Activo")
            cell_est.alignment = Alignment(horizontal="center")
            if es_novedad:
                cell_est.font = Font(name="Segoe UI", size=9, bold=True, color="059669")

            ws2.cell(row=row_idx, column=4, value=len(art.get("contenido_markdown", "")) or len(art.get("titulo", ""))).alignment = Alignment(horizontal="center")

            cell_url = ws2.cell(row=row_idx, column=5, value=art_url)
            if art_url and art_url.startswith("http"):
                cell_url.hyperlink = art_url
                cell_url.font = Font(name="Segoe UI", size=9, color="2563EB", underline="single")

            for c in range(1, 6):
                cell_i = ws2.cell(row=row_idx, column=c)
                if fill_actual:
                    cell_i.fill = fill_actual
                cell_i.border = border_thin
                if c != 3 and not (c == 5 and art_url.startswith("http")):
                    cell_i.font = font_normal

        # Habilitar autofiltro en tabla de datos
        ws2.auto_filter.ref = f"A1:E{len(lista_articulos) + 1}"

        ws2.column_dimensions["A"].width = 6
        ws2.column_dimensions["B"].width = 55
        ws2.column_dimensions["C"].width = 12
        ws2.column_dimensions["D"].width = 14
        ws2.column_dimensions["E"].width = 45

        # =========================================================
        # HOJA 3: MATRIZ DE INTELIGENCIA
        # =========================================================
        ws3 = wb.create_sheet(title="Matriz de Inteligencia")
        ws3.views.sheetView[0].showGridLines = True

        # Tabla de Categorías
        ws3["A1"] = "Categorías Temáticas"
        ws3["B1"] = "Menciones"
        ws3["C1"] = "Porcentaje"
        for col_idx, col_letter in enumerate(["A", "B", "C"], start=1):
            ws3[f"{col_letter}1"].font = font_header
            ws3[f"{col_letter}1"].fill = fill_slate
            ws3[f"{col_letter}1"].alignment = Alignment(horizontal="center")

        r3 = 2
        for cat in analisis_ia.get("categorias", []):
            ws3[f"A{r3}"] = cat.get("nombre", "")
            ws3[f"B{r3}"] = cat.get("cantidad", 0)
            ws3[f"C{r3}"] = f"{cat.get('porcentaje', 0)}%"
            ws3[f"B{r3}"].alignment = Alignment(horizontal="center")
            ws3[f"C{r3}"].alignment = Alignment(horizontal="center")
            for c_letter in ["A", "B", "C"]:
                ws3[f"{c_letter}{r3}"].font = font_normal
                ws3[f"{c_letter}{r3}"].border = border_thin
            r3 += 1

        # Tabla de Sentimiento
        r3 += 2
        ws3[f"A{r3}"] = "Tono / Sentimiento"
        ws3[f"B{r3}"] = "Frecuencia"
        ws3[f"C{r3}"] = "Proporción"
        for col_letter in ["A", "B", "C"]:
            ws3[f"{col_letter}{r3}"].font = font_header
            ws3[f"{col_letter}{r3}"].fill = fill_slate
            ws3[f"{col_letter}{r3}"].alignment = Alignment(horizontal="center")

        r3 += 1
        sents = analisis_ia.get("sentimientos", {})
        total_sents = sum(sents.values()) or 1
        for tono, cant in sents.items():
            ws3[f"A{r3}"] = str(tono).capitalize()
            ws3[f"B{r3}"] = cant
            ws3[f"C{r3}"] = f"{round((cant / total_sents) * 100, 1)}%"
            ws3[f"B{r3}"].alignment = Alignment(horizontal="center")
            ws3[f"C{r3}"].alignment = Alignment(horizontal="center")
            for c_letter in ["A", "B", "C"]:
                ws3[f"{c_letter}{r3}"].font = font_normal
                ws3[f"{c_letter}{r3}"].border = border_thin
            r3 += 1

        # Entidades Estructuradas
        r3 += 2
        ws3[f"A{r3}"] = "Clasificación de Entidades"
        ws3[f"B{r3}"] = "Elementos Identificados"
        ws3.merge_cells(f"B{r3}:E{r3}")
        for c_letter in ["A", "B", "C", "D", "E"]:
            ws3[f"{c_letter}{r3}"].font = font_header
            ws3[f"{c_letter}{r3}"].fill = fill_navy
            ws3[f"{c_letter}{r3}"].alignment = Alignment(horizontal="center")

        entidades = analisis_ia.get("entidades_estructuradas", {})
        mapeo_ent = [
            ("Instituciones y Empresas", entidades.get("instituciones_y_empresas", [])),
            ("Personas Relevantes", entidades.get("personas_relevantes", [])),
            ("Normas, Decretos y Leyes", entidades.get("marcos_legales_o_normas", []))
        ]

        r3 += 1
        for tipo_e, lista_e in mapeo_ent:
            ws3[f"A{r3}"] = tipo_e
            ws3[f"A{r3}"].font = font_bold
            ws3[f"A{r3}"].fill = fill_zebra
            ws3.merge_cells(f"B{r3}:E{r3}")
            ws3[f"B{r3}"] = ", ".join(lista_e) if lista_e else "Sin menciones explícitas."
            ws3[f"B{r3}"].font = font_normal
            for c_letter in ["A", "B", "C", "D", "E"]:
                ws3[f"{c_letter}{r3}"].border = border_thin
            r3 += 1

        # Recomendaciones
        r3 += 2
        ws3[f"A{r3}"] = "Recomendaciones Estratégicas y Próximos Pasos"
        ws3.merge_cells(f"A{r3}:E{r3}")
        ws3[f"A{r3}"].font = font_header
        ws3[f"A{r3}"].fill = fill_navy

        r3 += 1
        for rec in analisis_ia.get("recomendaciones_estrategicas", []):
            ws3[f"A{r3}"] = f"✔ {rec}"
            ws3.merge_cells(f"A{r3}:E{r3}")
            ws3[f"A{r3}"].font = font_normal
            for c_letter in ["A", "B", "C", "D", "E"]:
                ws3[f"{c_letter}{r3}"].border = border_thin
            r3 += 1

        ws3.column_dimensions["A"].width = 28
        ws3.column_dimensions["B"].width = 16
        ws3.column_dimensions["C"].width = 16
        ws3.column_dimensions["D"].width = 20
        ws3.column_dimensions["E"].width = 20

        # Guardar archivo
        output_file.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_file))
        return output_file
